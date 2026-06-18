"""
TC 결과를 xlsx 템플릿에 작성하는 모듈
"""
import shutil
from datetime import datetime

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from paths import resource

TEMPLATE_PATH = resource("sample.xlsx")

# 템플릿 열 위치 (1-indexed, 병합 범위의 최상단 좌측 셀 기준)
COL_NO = 3        # C
COL_대분류 = 4    # D
COL_중분류 = 5    # E
COL_소분류 = 6    # F
COL_확인항목 = 7  # G (G~J 수평 병합)
COL_결과 = 11     # K (K~L 수평 병합)
COL_JIRA = 13     # M (M~N 수평 병합)
COL_비고 = 15     # O (O~P 수평 병합)

# 각 데이터 행에서 복원해야 할 수평 병합 범위 (start_col, end_col)
_ROW_H_MERGES = [
    (COL_확인항목, 10),  # G~J
    (COL_결과,     12),  # K~L
    (COL_JIRA,     14),  # M~N
    (COL_비고,     16),  # O~P
]

DATA_START_ROW = 17


def _safe_set(ws, row: int, col: int, value):
    """병합 셀 내부(read-only)는 건너뛰고 top-left 셀에만 값 입력"""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _merge_column(ws, start_row: int, col: int, group_keys: list):
    """
    group_keys: 병합 그룹을 결정하는 키 리스트 (값 또는 튜플).
    연속된 동일 키를 가진 셀을 수직 병합.
    키를 tc_list에서 직접 가져오므로 MergedCell 읽기 문제가 없음.
    계층 키(튜플)를 사용하면 부모 그룹 경계에서 자동으로 병합이 끊김.
    """
    n = len(group_keys)
    if n == 0:
        return
    merge_start = 0
    for i in range(1, n + 1):
        is_new_group = (i == n) or (group_keys[i] != group_keys[i - 1])
        if is_new_group:
            actual_start = start_row + merge_start
            actual_end = start_row + i - 1
            if actual_end > actual_start:
                ws.merge_cells(
                    start_row=actual_start, start_column=col,
                    end_row=actual_end,     end_column=col,
                )
            ws.cell(row=actual_start, column=col).alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )
            merge_start = i


def write_tc(tc_list: list[dict], output_path: str, page_title: str = "") -> str:
    """
    TC 목록을 템플릿에 작성하고 output_path에 저장
    반환: 저장된 파일 경로
    """
    shutil.copy2(TEMPLATE_PATH, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # TEST Information - 일시 입력 (E7: E7:G8 병합 범위의 top-left)
    _safe_set(ws, 7, 5, datetime.now().strftime("%Y-%m-%d"))

    # DATA_START_ROW 이후의 기존 병합 셀 해제
    to_unmerge = [str(r) for r in ws.merged_cells.ranges if r.min_row >= DATA_START_ROW]
    for mr in to_unmerge:
        ws.unmerge_cells(mr)

    # 기존 데이터 초기화
    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col in [COL_NO, COL_대분류, COL_중분류, COL_소분류,
                    COL_확인항목, COL_결과, COL_JIRA, COL_비고]:
            ws.cell(row=row, column=col).value = None

    # TC 데이터 입력
    for i, tc in enumerate(tc_list):
        row = DATA_START_ROW + i
        _safe_set(ws, row, COL_NO,       i + 1)
        _safe_set(ws, row, COL_대분류,   tc.get("대분류", ""))
        _safe_set(ws, row, COL_중분류,   tc.get("중분류", ""))
        _safe_set(ws, row, COL_소분류,   tc.get("소분류", ""))
        _safe_set(ws, row, COL_확인항목, tc.get("확인 항목", ""))
        _safe_set(ws, row, COL_결과,     "Incomplete")
        _safe_set(ws, row, COL_비고,     tc.get("비고", ""))

        # 기본 정렬 (수평 병합 전에 적용)
        for col in [COL_NO, COL_대분류, COL_중분류, COL_소분류, COL_확인항목, COL_결과]:
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(wrap_text=True, vertical="center")

    if not tc_list:
        wb.save(output_path)
        return output_path

    n = len(tc_list)
    last_row = DATA_START_ROW + n - 1

    # ── 대/중/소분류 수직 병합 ──────────────────────────────────────────────
    # 계층 키를 tc_list에서 직접 구성 → 병합 도중 MergedCell 읽기 문제 없음
    # 튜플 키를 사용해 부모 그룹이 달라지면 자식도 자동으로 분리됨
    keys_대 = [tc.get("대분류", "") for tc in tc_list]
    keys_중 = [(tc.get("대분류", ""), tc.get("중분류", "")) for tc in tc_list]
    keys_소 = [(tc.get("대분류", ""), tc.get("중분류", ""), tc.get("소분류", "")) for tc in tc_list]

    _merge_column(ws, DATA_START_ROW, COL_대분류, keys_대)
    _merge_column(ws, DATA_START_ROW, COL_중분류, keys_중)
    _merge_column(ws, DATA_START_ROW, COL_소분류, keys_소)

    # ── 수평 병합 복원 (G~J, K~L, M~N, O~P) ────────────────────────────────
    for i in range(n):
        row = DATA_START_ROW + i
        for start_c, end_c in _ROW_H_MERGES:
            ws.merge_cells(start_row=row, start_column=start_c,
                           end_row=row,   end_column=end_c)
            ws.cell(row=row, column=start_c).alignment = Alignment(
                wrap_text=True, vertical="center"
            )

    # ── COUNTIF 수식 범위 업데이트 ──────────────────────────────────────────
    count_range = f"K{DATA_START_ROW}:K{last_row}"
    _safe_set(ws, 14, 6,  "=SUM(G14:L14)")
    _safe_set(ws, 14, 7,  f'=COUNTIF({count_range},"Incomplete")')
    _safe_set(ws, 14, 8,  f'=COUNTIF({count_range},"Pass")')
    _safe_set(ws, 14, 9,  f'=COUNTIF({count_range},"Fail")')
    _safe_set(ws, 14, 10, f'=COUNTIF({count_range},"N/A")')
    _safe_set(ws, 14, 11, f'=COUNTIF({count_range},"Block")')

    wb.save(output_path)
    return output_path
